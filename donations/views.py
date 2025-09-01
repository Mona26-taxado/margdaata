from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.contrib.auth import login, authenticate
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
import json
import uuid
import os
from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Import models
from .models import Customer, CustomUser, SahyogReceipt, BloodDonation, VyawasthaShulkReceipt, Sahyog, Notification
from django.db.models import Sum

# Import forms
from .forms import CustomUserCreationForm, SahyogForm, SahyogReceiptForm

from django.contrib.auth.backends import ModelBackend

#  Custom authentication backend that supports both email and username
class CustomAuthBackend(ModelBackend):
    def authenticate(self, request, username=None, password=None, **kwargs):
        #  Prevents crashes by returning None for missing credentials
        if username is None or password is None:
            return None

        try:
            if "@" in username:
                user = CustomUser.objects.get(email=username)
            else:
                user = CustomUser.objects.get(username=username)
        except CustomUser.DoesNotExist:
            return None  #  User not found, return None

        #  Ensures password verification is correct
        if user and user.check_password(password):
            return user

        return None

#  Login View (Supports Both Admin & User Login)
def custom_admin_login(request):
    if request.user.is_authenticated:
        return redirect("admin_dashboard" if request.user.role == "admin" else "user_dashboard")

    if request.method == "POST":
        email_or_username = request.POST.get("username")  #  Fixed: template uses "username" field
        password = request.POST.get("password")
        
        # Try to authenticate with email first
        user = authenticate(request, username=email_or_username, password=password)
        
        # If that fails, try with username
        if not user:
            try:
                user_obj = CustomUser.objects.get(username=email_or_username)
                user = authenticate(request, username=user_obj.username, password=password)
            except CustomUser.DoesNotExist:
                user = None

        if user:
            if user.is_approved and user.is_active:
                login(request, user)
                return redirect("admin_dashboard" if user.role == "admin" else "user_dashboard")
            else:
                messages.error(request, "Your account is pending approval or inactive.")
        else:
            messages.error(request, "Invalid email/username or password.")

    return render(request, "donation/login.html")

from django.contrib.auth.views import PasswordResetConfirmView, PasswordResetView
from django.contrib.auth.forms import SetPasswordForm
from django.urls import reverse_lazy
from django.contrib.auth import update_session_auth_hash
import logging

logger = logging.getLogger(__name__)

class CustomPasswordResetView(PasswordResetView):
    template_name = "donation/user_password_reset.html"
    email_template_name = "donation/password_reset_email.html"
    subject_template_name = "donation/password_reset_subject.txt"
    success_url = reverse_lazy("user_password_reset_done")
    from_email = settings.EMAIL_HOST_USER

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['domain'] = settings.SITE_DOMAIN
        context['protocol'] = 'https' if not settings.DEBUG else 'http'
        context['title'] = 'Password Reset'
        return context

    def form_valid(self, form):
        email = form.cleaned_data['email']
        logger.info(f"Password reset requested for email: {email}")
        
        try:
            user = CustomUser.objects.get(email=email)
            logger.info(f"User found with email: {email}")
            
            extra_context = {
                'domain': settings.SITE_DOMAIN,
                'protocol': 'https' if not settings.DEBUG else 'http',
                'email': email,
            }
            
            # Only send email if user exists
            form.save(
                request=self.request,
                from_email=self.from_email,
                email_template_name=self.email_template_name,
                subject_template_name=self.subject_template_name,
                extra_email_context=extra_context
            )
            logger.info(f"Password reset email sent successfully to {email}")
            return super().form_valid(form)
                
        except CustomUser.DoesNotExist:
            logger.warning(f"No user found with email: {email}")
            # Still return success to prevent email enumeration
            return super().form_valid(form)

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = "donation/user_password_reset_confirm.html"
    success_url = reverse_lazy("user_password_reset_complete")
    form_class = SetPasswordForm

    def form_valid(self, form):
        user = form.save(commit=False)  # Get user instance
        raw_password = form.cleaned_data["new_password1"]  # Capture new password
        user.set_password(raw_password)  # Hash and set password
        user.save()  # Force save
        update_session_auth_hash(self.request, user)  # Keep session active
        print(f"Password changed successfully for {user.email}")  # Debugging print
        return super().form_valid(form)

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

def send_html_email(to_email, subject, html_content):
    """Send HTML email with professional styling"""
    sender_email = "margdatatrust2025@gmail.com"
    sender_password = "gfsk zxli lmkv gygp"
    
    # Create message
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = to_email
    
    # Attach HTML content
    html_part = MIMEText(html_content, 'html')
    msg.attach(html_part)
    
    try:
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_email, msg.as_string())
        server.quit()
        print(f"HTML email sent successfully to {to_email}!")
    except Exception as e:
        print(f"Error sending HTML email: {e}")

def get_registration_notification_html(customer):
    """Generate HTML for new registration notification to admin"""
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background-color: #f4f4f4; }}
            .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 28px; font-weight: 300; }}
            .content {{ padding: 30px; }}
            .highlight {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; border-radius: 4px; }}
            .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin: 20px 0; }}
            .info-item {{ background: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 4px solid #007bff; }}
            .info-label {{ font-weight: bold; color: #495057; font-size: 12px; text-transform: uppercase; margin-bottom: 5px; }}
            .info-value {{ color: #212529; font-size: 16px; }}
            .cta-button {{ display: inline-block; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 12px 30px; text-decoration: none; border-radius: 25px; margin: 20px 0; font-weight: bold; }}
            .footer {{ background: #343a40; color: white; padding: 20px; text-align: center; font-size: 14px; }}
            .urgent {{ background: #f8d7da; border-left: 4px solid #dc3545; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🆕 New Registration Request</h1>
                <p>Margdata Trust - Admin Notification</p>
            </div>
            
            <div class="content">
                <div class="highlight urgent">
                    <strong>⚠️ Action Required:</strong> A new member has registered and requires your approval.
                </div>
                
                <h2 style="color: #495057; margin-bottom: 20px;">Registration Details</h2>
                
                <div class="info-grid">
                    <div class="info-item">
                        <div class="info-label">Full Name</div>
                        <div class="info-value">{customer.name}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">Email Address</div>
                        <div class="info-value">{customer.email}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">Mobile Number</div>
                        <div class="info-value">{customer.mobile}</div>
                    </div>
                    <div class="info-item">
                        <div class="info-label">Registration Date</div>
                        <div class="info-value">{customer.created_at.strftime("%d %B %Y at %I:%M %p")}</div>
                    </div>
                </div>
                
                <div class="highlight">
                    <strong>📋 Additional Information:</strong><br>
                    • Home District: {customer.home_district}<br>
                    • Department: {customer.department or 'Not specified'}<br>
                    • Reference: {customer.reference_name or 'Not specified'}
                </div>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="#" class="cta-button">Review & Approve</a>
                </div>
                
                <p style="color: #6c757d; font-size: 14px; text-align: center;">
                    Please log into your admin dashboard to review and approve/reject this registration request.
                </p>
            </div>
            
            <div class="footer">
                <p>© 2025 Margdata Trust. All rights reserved.</p>
                <p>This is an automated notification. Please do not reply to this email.</p>
            </div>
        </div>
    </body>
    </html>
    """
    return html

def get_approval_email_html(user, customer):
    """Generate HTML for user approval notification"""
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background-color: #f4f4f4; }}
            .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #28a745 0%, #20c997 100%); color: white; padding: 30px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 28px; font-weight: 300; }}
            .content {{ padding: 30px; }}
            .success-badge {{ background: #d4edda; color: #155724; padding: 10px 20px; border-radius: 25px; display: inline-block; font-weight: bold; margin: 20px 0; }}
            .welcome-box {{ background: linear-gradient(135deg, #e3f2fd 0%, #f3e5f5 100%); padding: 25px; border-radius: 10px; margin: 20px 0; text-align: center; }}
            .member-info {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #28a745; }}
            .cta-button {{ display: inline-block; background: linear-gradient(135deg, #28a745 0%, #20c997 100%); color: white; padding: 15px 35px; text-decoration: none; border-radius: 25px; margin: 20px 0; font-weight: bold; font-size: 16px; }}
            .footer {{ background: #343a40; color: white; padding: 20px; text-align: center; font-size: 14px; }}
            .features {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin: 20px 0; }}
            .feature {{ background: #fff; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #e9ecef; }}
            .feature-icon {{ font-size: 24px; margin-bottom: 10px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎉 Welcome to Margdata Trust!</h1>
                <p>Your registration has been approved</p>
            </div>
            
            <div class="content">
                <div class="success-badge">
                    ✅ ACCOUNT ACTIVATED
                </div>
                
                <div class="welcome-box">
                    <h2 style="color: #28a745; margin-bottom: 15px;">Congratulations, {customer.name}!</h2>
                    <p style="font-size: 18px; color: #495057; margin: 0;">
                        Your membership has been successfully approved. You are now an official member of Margdata Trust!
                    </p>
                </div>
                
                <div class="member-info">
                    <h3 style="color: #495057; margin-bottom: 15px;">Your Member Details</h3>
                    <p><strong>Member ID:</strong> {customer.md_code}</p>
                    <p><strong>Email:</strong> {user.email}</p>
                    <p><strong>Mobile:</strong> {customer.mobile}</p>
                    <p><strong>Registration Date:</strong> {customer.created_at.strftime("%d %B %Y")}</p>
                </div>
                
                <h3 style="color: #495057; margin: 20px 0;">What's Next?</h3>
                <div class="features">
                    <div class="feature">
                        <div class="feature-icon">🔐</div>
                        <strong>Login to Dashboard</strong>
                        <p style="font-size: 14px; color: #6c757d;">Access your personalized member dashboard</p>
                    </div>
                    <div class="feature">
                        <div class="feature-icon">📋</div>
                        <strong>Complete Profile</strong>
                        <p style="font-size: 14px; color: #6c757d;">Update your personal information</p>
                    </div>
                    <div class="feature">
                        <div class="feature-icon">💝</div>
                        <strong>Make Donations</strong>
                        <p style="font-size: 14px; color: #6c757d;">Contribute to various causes</p>
                    </div>
                    <div class="feature">
                        <div class="feature-icon">🆔</div>
                        <strong>Get ID Card</strong>
                        <p style="font-size: 14px; color: #6c757d;">Download your member ID card</p>
                    </div>
                </div>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="#" class="cta-button">Login to Dashboard</a>
                </div>
                
                <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; border-radius: 4px; margin: 20px 0;">
                    <strong>💡 Tip:</strong> Keep your login credentials safe. You can reset your password anytime from the login page.
                </div>
            </div>
            
            <div class="footer">
                <p>© 2025 Margdata Trust. All rights reserved.</p>
                <p>Thank you for joining our community of mutual support!</p>
            </div>
        </div>
    </body>
    </html>
    """
    return html

def get_rejection_email_html(user_email):
    """Generate HTML for user rejection notification"""
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background-color: #f4f4f4; }}
            .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #dc3545 0%, #fd7e14 100%); color: white; padding: 30px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 28px; font-weight: 300; }}
            .content {{ padding: 30px; }}
            .notice-box {{ background: #f8d7da; border-left: 4px solid #dc3545; padding: 20px; border-radius: 4px; margin: 20px 0; }}
            .info-box {{ background: #e2e3e5; padding: 20px; border-radius: 8px; margin: 20px 0; }}
            .contact-info {{ background: #d1ecf1; border-left: 4px solid #17a2b8; padding: 15px; border-radius: 4px; margin: 20px 0; }}
            .footer {{ background: #343a40; color: white; padding: 20px; text-align: center; font-size: 14px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>📋 Registration Update</h1>
                <p>Margdata Trust - Important Notice</p>
            </div>
            
            <div class="content">
                <div class="notice-box">
                    <h2 style="color: #721c24; margin-bottom: 15px;">Registration Status Update</h2>
                    <p style="font-size: 16px; color: #721c24; margin: 0;">
                        We regret to inform you that your registration request has been reviewed and could not be approved at this time.
                    </p>
                </div>
                
                <div class="info-box">
                    <h3 style="color: #495057; margin-bottom: 15px;">What This Means</h3>
                    <ul style="color: #495057; line-height: 1.6;">
                        <li>Your account has been removed from our system</li>
                        <li>You will need to submit a new registration if you wish to apply again</li>
                        <li>Please ensure all information provided is accurate and complete</li>
                    </ul>
                </div>
                
                <div class="contact-info">
                    <h3 style="color: #0c5460; margin-bottom: 10px;">Need Help?</h3>
                    <p style="color: #0c5460; margin: 0;">
                        If you have questions about this decision or would like to discuss your application, 
                        please contact our support team. We're here to help!
                    </p>
                </div>
                
                <div style="text-align: center; margin: 30px 0;">
                    <p style="color: #6c757d; font-size: 14px;">
                        Thank you for your interest in Margdata Trust.
                    </p>
                </div>
            </div>
            
            <div class="footer">
                <p>© 2025 Margdata Trust. All rights reserved.</p>
                <p>This is an automated notification. Please do not reply to this email.</p>
            </div>
        </div>
    </body>
    </html>
    """
    return html

from django.contrib.auth.hashers import check_password, make_password
import fitz  # PyMuPDF
from django.http import FileResponse, HttpResponse
from io import BytesIO
import qrcode
import logging

from django.contrib.auth import get_user_model
import os

from django.shortcuts import render
from django.db.models import Sum
from django.core.paginator import Paginator
from django.db.models import Q
from django.urls import reverse
from django.views.decorators.http import require_POST

def admin_dashboard(request):
    total_users = CustomUser.objects.count()
    total_sahyog = SahyogReceipt.objects.count() # Changed from Sahyog to SahyogReceipt
    total_blood_donors = BloodDonation.objects.count()
    total_vyawastha_shulk = VyawasthaShulkReceipt.objects.aggregate(total=Sum('amount'))['total'] or 0

    context = {
        "total_users": total_users,
        "total_sahyog": total_sahyog,
        "total_blood_donors": total_blood_donors,
        "total_vyawastha_shulk": total_vyawastha_shulk,
    }

    return render(request, "donation/admin_dashboard.html", context)

CustomUser = get_user_model()

def login_view(request):
    if request.method == "POST":
        email = request.POST.get("username")  #  Users enter email, but Django expects 'username'
        password = request.POST.get("password")

        print(f" Debug: Login Email = {email}, Password = {password}")

        #  Find user by email
        user = CustomUser.objects.filter(email=email).first()

        if not user:
            print(" Debug: No user found with this email")
            return render(request, "donation/login.html", {"error": "Invalid email or password"})

        print(f" Debug: Found user with email = {user.email}")

        #  Authenticate using 'username' argument (Django expects username, not email)
        user = authenticate(request, username=email, password=password)

        if user is None:
            print(" Debug: Authentication failed, incorrect password")
            return render(request, "donation/login.html", {"error": "Invalid email or password"})

        print(f" Debug: Authentication successful for {user.email}")

        login(request, user)

        return redirect("admin_dashboard" if user.is_staff else "user_dashboard")

    return render(request, "donation/login.html")

from django.contrib import messages  #  Import messages for Bootstrap alerts

def register_customer(request):
    if request.method == "POST":
        form = CustomerRegistrationForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                cleaned_data = form.cleaned_data
                email = cleaned_data.get("email")
                password = cleaned_data.get("password")  # Get password from form

                # Check for duplicate email
                if CustomUser.objects.filter(email=email).exists():
                    error_msg = "This email is already registered. Please use a different email or log in."
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'errors': {'email': [error_msg]}}, status=400)
                    else:
                        form.add_error('email', error_msg)
                        return render(request, "donation/register.html", {"form": form})

                if not password:
                    error_msg = "Password is required."
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'errors': {'password': [error_msg]}}, status=400)
                    else:
                        form.add_error('password', error_msg)
                        return render(request, "donation/register.html", {"form": form})

                user = CustomUser.objects.create(
                    email=email,
                    role="user",
                    is_approved=False,
                    is_active=False
                )
                user.set_password(password)
                user.save()

                customer = form.save(commit=False)
                customer.user = user
                customer.approved = False
                customer.save()

                # Generate ID card automatically for new user
                generate_user_id_card(customer)

                # Send notification email to admins
                admin_emails = CustomUser.objects.filter(is_staff=True).values_list('email', flat=True)
                if admin_emails:
                    # Send HTML email to each admin
                    for admin_email in admin_emails:
                        html_content = get_registration_notification_html(customer)
                        send_html_email(admin_email, '🆕 New Registration Request - Action Required', html_content)

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'redirect_url': reverse('registration_success')})
                messages.success(request, "Registration successful! Await admin approval.")
                return redirect('registration_success')
            else:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'errors': form.errors}, status=400)
                else:
                    return render(request, "donation/register.html", {"form": form})
        except Exception as e:
            logging.exception("Registration error")
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': {'__all__': [str(e)]}}, status=500)
            else:
                messages.error(request, f"Error: {str(e)}")
                return render(request, "donation/register.html", {"form": form})
    else:
        form = CustomerRegistrationForm()
    return render(request, "donation/register.html", {"form": form})

@login_required
def approve_user(request, user_id):
    if request.method != "POST":  #  Rejects GET requests
        return JsonResponse({"error": "Invalid request method"}, status=405)

    #  Ensure the user exists
    user = get_object_or_404(CustomUser, id=user_id)

    #  Ensure the customer exists
    customer = Customer.objects.filter(user=user).first()
    if not customer:
        return JsonResponse({"error": "No customer found"}, status=404)

    #  Approve and activate the user
    user.is_approved = True
    user.is_active = True
    user.save()

    #  Approve the customer entry
    customer.approved = True
    customer.save()

    #  Send Approval Email Notification
    html_content = get_approval_email_html(user, customer)
    send_html_email(user.email, '🎉 Welcome to Margdata Trust - Account Approved!', html_content)

    return redirect("manage_members")

from django.urls import reverse

def reject_user(request, user_id):
    if request.method == "POST":
        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            messages.error(request, " User not found or already deleted.")
            return redirect(reverse("manage_members"))

        user_email = user.email  # Save email before deleting
        user.delete()  #  Completely remove user

        #  Send rejection email
        html_content = get_rejection_email_html(user_email)
        send_html_email(user_email, '📋 Registration Status Update - Important Notice', html_content)

        messages.success(request, f" User {user_email} has been rejected and deleted.")
    
    return redirect(reverse("manage_members"))  # Redirect back to admin panel

from django.contrib.auth.backends import ModelBackend

#  Custom authentication backend that supports both email and username
class CustomAuthBackend(ModelBackend):
    def authenticate(self, request, username=None, password=None, **kwargs):
        #  Prevents crashes by returning None for missing credentials
        if username is None or password is None:
            return None

        try:
            if "@" in username:
                user = CustomUser.objects.get(email=username)
            else:
                user = CustomUser.objects.get(username=username)
        except CustomUser.DoesNotExist:
            return None  #  User not found, return None

        #  Ensures password verification is correct
        if user and user.check_password(password):
            return user

        return None

from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

@csrf_exempt  # Bypasses CSRF token issues (for GET logout)
def user_logout(request):
    if request.method == "POST" or request.method == "GET":  # Accept both methods
        logout(request)
        return redirect("login")  # Redirect to login page

def generate_payment_receipt(request, customer_id):
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        return HttpResponse("Customer not found", status=404)

    # Set up the response as a PDF file
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Payment_Receipt_{customer.md_code}.pdf"'

    # Create the PDF object
    pdf = canvas.Canvas(response, pagesize=letter)
    pdf.setTitle(f"Payment Receipt - {customer.md_code}")

    # Add content to the PDF
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(200, 750, "Payment Receipt")
    pdf.setFont("Helvetica", 12)

    pdf.drawString(100, 700, f"MD Code: {customer.md_code}")
    pdf.drawString(100, 680, f"Name: {customer.name}")
    pdf.drawString(100, 660, f"Email: {customer.email}")
    pdf.drawString(100, 640, f"Mobile: {customer.mobile}")
    pdf.drawString(100, 620, f"Nominee: {customer.nominee_name} ({customer.nominee_relation})")

    pdf.drawString(100, 570, "Thank you for your payment!")

    pdf.showPage()
    pdf.save()

    return response

#  Admin Dashboard (Only for Custom Admin)
@login_required
def custom_admin_dashboard(request):
    if request.user.role != "admin":
        return redirect('user_dashboard')  # If user is not admin, redirect to user dashboard
    return render(request, "donation/admin_dashboard.html")

#  User Dashboard (For Normal Users)
@login_required
def user_dashboard(request):
    if request.user.is_staff:
        return HttpResponseForbidden("You are not allowed to access this page.")

    # Get customer data for the logged-in user
    try:
        customer = Customer.objects.get(user=request.user)
        user_name = customer.name  # Get the customer's name
    except Customer.DoesNotExist:
        customer = None
        user_name = request.user.email  # Fallback to email if no customer profile

    notifications = Notification.objects.filter(
        Q(target_type="all_users") |
        Q(target_type="specific_user", specific_user=request.user)
    ).order_by('-created_at')

    return render(request, 'donation/user_dashboard.html', {
        'notifications': notifications,
        'customer': customer,
        'user_name': user_name  # Pass user name to template
    })

def manage_members(request):
    # Get search query
    search_query = request.GET.get('search', '')
    
    # Fetch approved users with search filter
    approved_users = Customer.objects.filter(approved=True)
    if search_query:
        approved_users = approved_users.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(mobile__icontains=search_query)
        )
    approved_users = approved_users.order_by("-created_at")
    
    # Fetch pending requests with search filter
    pending_requests = Customer.objects.filter(approved=False)
    if search_query:
        pending_requests = pending_requests.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(mobile__icontains=search_query)
        )
    pending_requests = pending_requests.order_by("-created_at")

    # Pagination for approved users (10 per page)
    approved_paginator = Paginator(approved_users, 10)
    approved_page_number = request.GET.get('approved_page', 1)
    approved_page_obj = approved_paginator.get_page(approved_page_number)

    # Add serial numbers for approved users
    approved_start_index = (approved_page_obj.number - 1) * approved_paginator.per_page + 1
    for index, user in enumerate(approved_page_obj, start=approved_start_index):
        user.serial_number = index

    # Pagination for pending requests (10 per page)
    pending_paginator = Paginator(pending_requests, 10)
    pending_page_number = request.GET.get('pending_page', 1)
    pending_page_obj = pending_paginator.get_page(pending_page_number)

    # Add serial numbers for pending requests
    pending_start_index = (pending_page_obj.number - 1) * pending_paginator.per_page + 1
    for index, user in enumerate(pending_page_obj, start=pending_start_index):
        user.serial_number = index

    return render(request, "donation/manage_members.html", {
        "approved_users": approved_page_obj,
        "pending_requests": pending_page_obj,
        "search_query": search_query,
        "approved_paginator": approved_paginator,
        "pending_paginator": pending_paginator
    })

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Customer
from .forms import CustomerRegistrationForm  # Use the existing form

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Customer
from .forms import CustomerEditForm

# View Members: Separate for Admin and Users
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Customer  # Import your model

def view_members(request):
    search_query = request.GET.get('search', '')  # Get search query from request
    approved_users = Customer.objects.filter(approved=True)  # Fetch only approved users

    # If search query exists, filter by name, email, or mobile
    if search_query:
        approved_users = approved_users.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(mobile__icontains=search_query)
        )

    approved_users = approved_users.order_by("-created_at")  # Order by newest first

    # Pagination (10 users per page)
    paginator = Paginator(approved_users, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Add serial numbers based on page number
    start_index = (page_obj.number - 1) * paginator.per_page + 1
    for index, user in enumerate(page_obj, start=start_index):
        user.serial_number = index

    # Check if user is admin or normal user
    if request.user.is_staff:  
        return render(request, "donation/view_members.html", {
            "approved_users": page_obj,
            "search_query": search_query,
            "paginator": paginator,
            "page_obj": page_obj
        })
    else:  
        return render(request, "donation/user_view_members.html", {
            "approved_users": page_obj,
            "search_query": search_query,
            "paginator": paginator,
            "page_obj": page_obj
        })

# Edit Member Details (Only for Admin)
def edit_member(request, user_id):
    if not request.user.is_staff:  # Prevent users from editing
        messages.error(request, "Unauthorized Access!")
        return redirect("view_members")

    member = get_object_or_404(Customer, id=user_id)
    if request.method == "POST":
        form = CustomerEditForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, f"{member.name}'s details have been updated!")
            return redirect("view_members")
    else:
        form = CustomerEditForm(instance=member)

    return render(request, "donation/edit_member.html", {"form": form, "member": member})

# Delete Member (Only for Admin)
def delete_member(request, user_id):
    if not request.user.is_staff:  # Prevent users from deleting
        messages.error(request, "Unauthorized Access!")
        return redirect("view_members")

    member = get_object_or_404(Customer, id=user_id)
    member.delete()
    messages.error(request, f"{member.name} has been removed!")
    return redirect("view_members")

def user_view_members(request):
    search_query = request.GET.get('search', '')
    approved_users = Customer.objects.filter(approved=True).order_by("-created_at")
    if search_query:
        approved_users = approved_users.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(mobile__icontains=search_query)
        )
    paginator = Paginator(approved_users, 10)  # 10 per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Calculate starting index for serial numbers
    start_index = (page_obj.number - 1) * paginator.per_page + 1
    for index, user in enumerate(page_obj, start=start_index):
        user.serial_number = index
        
    return render(request, "donation/user_view_members.html", {
        "approved_users": page_obj,
        "search_query": search_query,
        "page_obj": page_obj,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_member(request):
    if request.method == "POST":
        print(" DEBUG: Received POST request to add member")
        form = CustomUserCreationForm(request.POST)

        if form.is_valid():
            try:
                user = form.save(commit=False)
                user.is_approved = True  #  Auto-approve admin-added users
                user.is_active = True  #  Activate user
                user.role = "user"  #  Ensure correct role
                
                print(f" DEBUG: Trying to save user {user.email}")  # Debug print
                
                user.save()
                
                print(f" User {user.email} saved successfully!")  # Debug print

                #  Ensure a `Customer` entry is created
                customer, created = Customer.objects.get_or_create(
                    user=user,
                    defaults={
                        "name": request.POST.get("name"),
                        "email": user.email,
                        "mobile": request.POST.get("mobile"),
                        "approved": True,  #  Auto-approve
                        "posting_state": request.POST.get("posting_state"),  #  Save State
                        "posting_district": request.POST.get("posting_district"),  #  Save District
                    }
                )

                print(f" Customer entry created for {user.email}")  # Debug print
                print(f" State: {customer.posting_state}, District: {customer.posting_district}")  # Debugging output

                messages.success(request, "New member added successfully.")
                return redirect('manage_members')

            except Exception as e:
                print(f" ERROR while saving user: {str(e)}")  # Print error message
                messages.error(request, f"Error adding member: {str(e)}")
        else:
            print(" FORM ERROR:", form.errors)  # Print validation errors
            messages.error(request, f"Form errors: {form.errors}")

    else:
        form = CustomUserCreationForm()

    return render(request, 'donation/add_member.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_staff)
def running_sahyog(request):
    if request.method == "POST":
        title = request.POST.get("title")
        account_holder_name = request.POST.get("account_holder_name")  #  New Field
        bank_name = request.POST.get("bank_name")
        account_number = request.POST.get("account_number")
        ifsc_code = request.POST.get("ifsc_code")
        qr_code = request.FILES.get("qr_code")

        sahyog = Sahyog(
            title=title,
            account_holder_name=account_holder_name,  #  Saving New Field
            bank_name=bank_name,
            account_number=account_number,
            ifsc_code=ifsc_code,
            qr_code=qr_code
        )
        sahyog.save()
        return redirect('sahyog_list')

    return render(request, "donation/running_sahyog.html")

#  Restrict Editing to Custom Admins Only
@login_required
@user_passes_test(lambda u: u.is_staff)  #  Only Custom Admin can access
def edit_sahyog(request, sahyog_id):
    sahyog = get_object_or_404(Sahyog, id=sahyog_id)  # Get Sahyog by ID

    if request.method == "POST":
        form = SahyogForm(request.POST, request.FILES, instance=sahyog)
        if form.is_valid():
            form.save()
            return redirect('sahyog_list')  # Redirect to Sahyog list after edit
    else:
        form = SahyogForm(instance=sahyog)

    return render(request, "donation/edit_sahyog.html", {"form": form, "sahyog": sahyog})

#  Restrict Deleting to Custom Admins Only
@login_required
@user_passes_test(lambda u: u.is_staff)  # Only admins can delete
def delete_sahyog(request, sahyog_id):
    sahyog = get_object_or_404(Sahyog, id=sahyog_id)
    sahyog.delete()
    return redirect('sahyog_list')

def sahyog_list(request):
    sahyog_entries = Sahyog.objects.all()  #  Get all Sahyog entries
    
    #  Choose base template dynamically
    base_template = "donation/base.html" if request.user.is_staff else "donation/base_user.html"
    
    return render(request, "donation/sahyog_list.html", {
        "sahyog_list": sahyog_entries,
        "base_template": base_template
    })

# Check if the user is a Custom Admin
def is_custom_admin(user):
    return user.is_staff  # Modify if you have a separate admin role

User = get_user_model()  #  This ensures it fetches `CustomUser`

def send_notification(request):
    users = User.objects.all()
    if request.method == "POST":
        form = NotificationForm(request.POST)
        if form.is_valid():
            target_type = form.cleaned_data["target_type"]
            specific_user_id = request.POST.get("specific_user")

            if target_type == "specific_user":
                if not specific_user_id:
                    messages.error(request, "Please select a user.")
                else:
                    notification = form.save(commit=False)
                    notification.specific_user_id = specific_user_id
                    notification.save()
                    messages.success(request, "Notification sent successfully!")
                    return redirect("send_notification")
            else:
                # For all users, clear specific_user
                notification = form.save(commit=False)
                notification.specific_user = None
                notification.save()
                messages.success(request, "Notification sent successfully!")
                return redirect("send_notification")
    else:
        form = NotificationForm()
    return render(request, "donation/send_notification.html", {"form": form, "users": users})

#  View to display notifications for admin
@login_required
@user_passes_test(is_custom_admin)
def manage_notifications(request):
    notifications = Notification.objects.order_by("-created_at")  # Show latest notifications first
    return render(request, "donation/manage_notifications.html", {"notifications": notifications})

@staff_member_required
def delete_notification(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    notification.delete()
    return redirect("manage_notifications")  # Redirect back to notifications list

@login_required
def user_profile(request):
    customer = get_object_or_404(Customer, user=request.user)

    if request.method == "POST":
        # Retrieve each field from request.POST
        customer.name = request.POST.get("name")
        customer.mobile = request.POST.get("mobile")
        customer.email = request.POST.get("email")
        customer.dob = request.POST.get("dob")
        customer.gender = request.POST.get("gender")
        customer.aadhar = request.POST.get("aadhar")
        customer.home_address = request.POST.get("home_address")
        customer.home_state = request.POST.get("home_state")
        customer.home_district = request.POST.get("home_district")
        customer.department = request.POST.get("department")
        customer.post = request.POST.get("post")
        customer.posting_state = request.POST.get("posting_state")
        customer.posting_district = request.POST.get("posting_district")
        customer.blood_group = request.POST.get("blood_group")
        customer.disease = request.POST.get("disease")

        # Nominee fields as well
        customer.first_nominee_name = request.POST.get("first_nominee_name")
        customer.first_nominee_relation = request.POST.get("first_nominee_relation")
        customer.first_nominee_mobile = request.POST.get("first_nominee_mobile")

        # Handle photo upload
        if 'photo' in request.FILES:
            customer.photo = request.FILES['photo']

        # Finally save
        customer.save()
        
        # Regenerate ID card with updated information
        generate_user_id_card(customer)
        
        messages.success(request, "Profile updated successfully. Your ID card has been updated.")
        return redirect("user_profile")  # refreshes the page

    # GET request just shows the form
    return render(request, "donation/user_profile.html", {"customer": customer})

@login_required
def user_download_id_card(request):
    """User view to download their own ID card"""
    customer = get_object_or_404(Customer, user=request.user)
    
    # Generate ID card if it doesn't exist
    filename = f"ID_Card_{customer.md_code}.pdf"
    filepath = os.path.join(settings.MEDIA_ROOT, 'id_cards', filename)
    
    # If file doesn't exist, generate it
    if not os.path.exists(filepath):
        generate_user_id_card(customer)
    
    # Check if file exists after generation
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    else:
        messages.error(request, 'Failed to generate ID card. Please try again.')
        return redirect('user_profile')

@login_required
def upload_receipt(request):
    if request.method == "POST":
        form = SahyogReceiptForm(request.POST, request.FILES)
        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.user = request.user  # Attach the receipt to the logged-in user
            receipt.sahyog = None  #  Ensure no Sahyog is assigned
            receipt.save()
            return redirect("user_receipts")  # Redirect to receipt list

    else:
        form = SahyogReceiptForm()

    return render(request, "donation/upload_receipt.html", {"form": form})

@login_required
def user_receipts(request):
    receipts = SahyogReceipt.objects.filter(user=request.user)
    return render(request, "donation/user_receipts.html", {"receipts": receipts})

@staff_member_required
def admin_receipts(request):
    receipts = SahyogReceipt.objects.all()
    return render(request, "donation/admin_receipts.html", {"receipts": receipts})

@login_required
def upload_vyawastha_shulk(request):
    if request.method == "POST":
        form = VyawasthaShulkReceiptForm(request.POST, request.FILES)
        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.user = request.user
            receipt.save()
            return redirect("user_vyawastha_shulk_receipts")  # Redirect to receipt list

    else:
        form = VyawasthaShulkReceiptForm()

    return render(request, "donation/upload_vyawastha_shulk.html", {"form": form})

@login_required
def user_vyawastha_shulk_receipts(request):
    receipts = VyawasthaShulkReceipt.objects.filter(user=request.user)
    return render(request, "donation/user_vyawastha_shulk_receipts.html", {"receipts": receipts})

@staff_member_required
def admin_vyawastha_shulk_receipts(request):
    receipts = VyawasthaShulkReceipt.objects.all()
    return render(request, "donation/admin_vyawastha_shulk_receipts.html", {"receipts": receipts})

@login_required
def submit_blood_donation(request):
    if request.method == "POST":
        form = BloodDonationForm(request.POST, request.FILES)
        if form.is_valid():
            blood_donation = form.save(commit=False)
            blood_donation.user = request.user
            blood_donation.save()
            return redirect("blood_donation_list")
    else:
        form = BloodDonationForm()
    return render(request, "donation/blood_donation_form.html", {"form": form})

@login_required
def blood_donation_list(request):
    donations = BloodDonation.objects.all().order_by("-created_at")
    
    # Use different base templates based on user type
    if request.user.is_staff:
        base_template = "donation/base.html"
    else:
        base_template = "donation/base_user.html"

    return render(request, "donation/blood_donation_list.html", {
        "donations": donations,
        "base_template": base_template
    })

@staff_member_required
def update_blood_donation_status(request, donation_id, status):
    donation = get_object_or_404(BloodDonation, id=donation_id)
    if status in ["Pending", "Approved", "Completed"]:
        donation.status = status
        donation.save()
    return redirect("blood_donation_list")

@login_required
def generate_id_card(request):
    try:
        customer = Customer.objects.get(user=request.user)
    except Customer.DoesNotExist:
        return HttpResponse("Customer profile not found", status=404)

    # Generate ID card using the new function
    if generate_user_id_card(customer):
        # Download the generated file
        filename = f"ID_Card_{customer.md_code}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, 'id_cards', filename)
        
        if os.path.exists(filepath):
            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        else:
            return HttpResponse("Failed to generate ID card", status=500)
    else:
        return HttpResponse("Failed to generate ID card", status=500)

#  Improved Text Wrapping Function
def add_wrapped_text(page, label, value, position, max_width=120, line_spacing=8):
    """Handles wrapping text if it exceeds max width."""
    words = (f"{label} {value}").split()
    current_line = ""
    y_offset = 0
    font_size = 8
    text_color = (0, 0, 0)
    font_regular = "times-roman"

    for word in words:
        test_line = f"{current_line} {word}".strip()
        text_width = page.insert_text(position, test_line, fontsize=font_size, fontname=font_regular, color=text_color)

        #  If text exceeds max_width, move to the next line
        if text_width > max_width:
            page.insert_text((position[0], position[1] + y_offset), current_line, fontsize=font_size, fontname=font_regular, color=text_color)
            current_line = word  # Start new line
            y_offset += line_spacing  # Adjust line spacing
        else:
            current_line = test_line

    #  Insert the last line
    page.insert_text((position[0], position[1] + y_offset), current_line, fontsize=font_size, fontname=font_regular, color=text_color)

import smtplib
from email.mime.text import MIMEText

def send_approval_email(user_email, username):
    sender_email = "margdatatrust2025@gmail.com"
    sender_password = "gfsk zxli lmkv gygp"

    subject = "Your ID is Activated"
    body = f"Hello {username},\n\nYour ID has been approved and is now active. You can log in and use our services.\n\nBest Regards,\nMargdata Team"

    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = sender_email
    msg["To"] = user_email

    try:
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)  # Change SMTP settings if needed
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, user_email, msg.as_string())
        server.quit()
        print("Approval email sent successfully!")
    except Exception as e:
        print("Error sending email:", e)

from django.contrib.auth.views import PasswordResetConfirmView, PasswordResetView
from django.contrib.auth.forms import SetPasswordForm
from django.urls import reverse_lazy
from django.contrib.auth import update_session_auth_hash
import logging

logger = logging.getLogger(__name__)

class CustomPasswordResetView(PasswordResetView):
    template_name = "donation/user_password_reset.html"
    email_template_name = "donation/password_reset_email.html"
    subject_template_name = "donation/password_reset_subject.txt"
    success_url = reverse_lazy("user_password_reset_done")
    from_email = settings.EMAIL_HOST_USER

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['domain'] = settings.SITE_DOMAIN
        context['protocol'] = 'https' if not settings.DEBUG else 'http'
        context['title'] = 'Password Reset'
        return context

    def form_valid(self, form):
        email = form.cleaned_data['email']
        logger.info(f"Password reset requested for email: {email}")
        
        try:
            user = CustomUser.objects.get(email=email)
            logger.info(f"User found with email: {email}")
            
            extra_context = {
                'domain': settings.SITE_DOMAIN,
                'protocol': 'https' if not settings.DEBUG else 'http',
                'email': email,
            }
            
            # Only send email if user exists
            form.save(
                request=self.request,
                from_email=self.from_email,
                email_template_name=self.email_template_name,
                subject_template_name=self.subject_template_name,
                extra_email_context=extra_context
            )
            logger.info(f"Password reset email sent successfully to {email}")
            return super().form_valid(form)
                
        except CustomUser.DoesNotExist:
            logger.warning(f"No user found with email: {email}")
            # Still return success to prevent email enumeration
            return super().form_valid(form)

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = "donation/user_password_reset_confirm.html"
    success_url = reverse_lazy("user_password_reset_complete")
    form_class = SetPasswordForm

    def form_valid(self, form):
        user = form.save(commit=False)  # Get user instance
        raw_password = form.cleaned_data["new_password1"]  # Capture new password
        user.set_password(raw_password)  # Hash and set password
        user.save()  # Force save
        update_session_auth_hash(self.request, user)  # Keep session active
        print(f"Password changed successfully for {user.email}")  # Debugging print
        return super().form_valid(form)

def upi_qr(request):
    upi_url = "upi://pay?pa=margdatatrust@sbi&pn=Margdata Trust&am=50&cu=INR"
    img = qrcode.make(upi_url)
    response = HttpResponse(content_type="image/png")
    img.save(response, "PNG")
    return response

@staff_member_required
def download_members_excel(request):
    """Download all approved members data as Excel file"""
    
    # Get all approved members
    members = Customer.objects.filter(approved=True).select_related('user')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Members Data"
    
    # Define headers
    headers = [
        'S.No', 'Name', 'Email', 'Mobile', 'Home Mobile', 'Date of Birth', 'Gender',
        'Aadhar Number', 'Home Address', 'Home State', 'Home District', 'Reference Name',
        'Sector', 'Department/Post', 'Posting State', 'Posting District',
        'Blood Group', 'Disease', 'Nominee Name', 'Nominee Relation', 'Nominee Mobile',
        'Registration Date', 'MD Code'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, member in enumerate(members, 2):
        ws.cell(row=row, column=1, value=row-1)  # S.No
        ws.cell(row=row, column=2, value=member.name)
        ws.cell(row=row, column=3, value=member.email)
        ws.cell(row=row, column=4, value=member.mobile)
        ws.cell(row=row, column=5, value=member.mobile_home or '')
        ws.cell(row=row, column=6, value=member.dob.strftime('%Y-%m-%d') if member.dob else '')
        ws.cell(row=row, column=7, value=member.gender)
        ws.cell(row=row, column=8, value=member.aadhar or '')
        ws.cell(row=row, column=9, value=member.home_address)
        ws.cell(row=row, column=10, value=member.home_state or '')
        ws.cell(row=row, column=11, value=member.home_district)
        ws.cell(row=row, column=12, value=member.reference_name or '')
        ws.cell(row=row, column=13, value=member.department or '')
        ws.cell(row=row, column=14, value=member.post or '')
        ws.cell(row=row, column=15, value=member.posting_state or '')
        ws.cell(row=row, column=16, value=member.posting_district or '')
        ws.cell(row=row, column=17, value=member.blood_group or '')
        ws.cell(row=row, column=18, value=member.disease or '')
        ws.cell(row=row, column=19, value=member.first_nominee_name or '')
        ws.cell(row=row, column=20, value=member.first_nominee_relation or '')
        ws.cell(row=row, column=21, value=member.first_nominee_mobile or '')
        ws.cell(row=row, column=22, value=member.user.date_joined.strftime('%Y-%m-%d %H:%M') if member.user else '')
        ws.cell(row=row, column=23, value=member.md_code)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="members_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

@staff_member_required
def download_sahyog_excel(request):
    """Download all sahyog data as Excel file"""
    
    # Get all sahyog receipts
    sahyog_receipts = SahyogReceipt.objects.all().select_related('user', 'sahyog')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sahyog Data"
    
    # Define headers
    headers = [
        'S.No', 'User Name', 'User Email', 'User Mobile', 'Sahyog Title',
        'Receipt Image', 'Uploaded Date', 'User Registration Date'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, receipt in enumerate(sahyog_receipts, 2):
        # Get customer details
        try:
            customer = Customer.objects.get(user=receipt.user)
            user_name = customer.name
            user_mobile = customer.mobile
        except Customer.DoesNotExist:
            user_name = receipt.user.email
            user_mobile = 'N/A'
        
        ws.cell(row=row, column=1, value=row-1)  # S.No
        ws.cell(row=row, column=2, value=user_name)
        ws.cell(row=row, column=3, value=receipt.user.email)
        ws.cell(row=row, column=4, value=user_mobile)
        ws.cell(row=row, column=5, value=receipt.sahyog.title if receipt.sahyog else 'N/A')
        ws.cell(row=row, column=6, value=str(receipt.receipt_image) if receipt.receipt_image else 'N/A')
        ws.cell(row=row, column=7, value=receipt.uploaded_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=8, value=receipt.user.date_joined.strftime('%Y-%m-%d %H:%M'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sahyog_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response




@staff_member_required
def admin_id_card_management(request):
    """Admin view to manage all user ID cards"""
    search_query = request.GET.get('search', '')
    customers = Customer.objects.all().select_related('user').order_by('-created_at')
    
    # If search query exists, filter by name, email, or mobile
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(mobile__icontains=search_query)
        )
    
    # Pagination (10 users per page)
    paginator = Paginator(customers, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Add serial numbers based on page number
    start_index = (page_obj.number - 1) * paginator.per_page + 1
    for index, customer in enumerate(page_obj, start=start_index):
        customer.serial_number = index
    
    context = {
        'customers': page_obj,
        'search_query': search_query,
        'paginator': paginator,
        'page_obj': page_obj,
        'title': 'User ID Card Management'
    }
    return render(request, 'donation/admin_id_card_management.html', context)

@staff_member_required
def admin_edit_id_card(request, customer_id):
    """Admin view to edit a specific user's ID card details"""
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        messages.error(request, 'Customer not found.')
        return redirect('admin_id_card_management')
    
    if request.method == 'POST':
        # Handle form submission
        customer.name = request.POST.get('name', customer.name)
        customer.email = request.POST.get('email', customer.email)
        customer.mobile = request.POST.get('mobile', customer.mobile)
        customer.blood_group = request.POST.get('blood_group', customer.blood_group)
        customer.home_state = request.POST.get('home_state', customer.home_state)
        customer.home_district = request.POST.get('home_district', customer.home_district)
        
        # Handle photo upload
        if 'photo' in request.FILES:
            customer.photo = request.FILES['photo']
        
        customer.save()
        
        # Generate new ID card automatically
        generate_user_id_card(customer)
        
        messages.success(request, f'ID card updated successfully for {customer.name}')
        return redirect('admin_id_card_management')
    
    context = {
        'customer': customer,
        'title': f'Edit ID Card - {customer.name}'
    }
    return render(request, 'donation/admin_edit_id_card.html', context)

@staff_member_required
def admin_generate_id_card(request, customer_id):
    """Admin view to manually generate ID card for a user"""
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        messages.error(request, 'Customer not found.')
        return redirect('admin_id_card_management')
    
    # Generate ID card
    generate_user_id_card(customer)
    
    messages.success(request, f'ID card generated successfully for {customer.name}')
    return redirect('admin_id_card_management')

def generate_user_id_card(customer):
    """Helper function to generate ID card for a user using the uploaded template"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.units import inch, cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        import os
        import math
        
        # Register good quality fonts for English text
        good_fonts_available = False
        try:
            import glob
            import platform
            
            if platform.system() == "Windows":
                # Try to use Calibri (modern, clean font)
                calibri_paths = glob.glob("C:/Windows/Fonts/calibri*.ttf")
                calibrib_paths = glob.glob("C:/Windows/Fonts/calibrib*.ttf")
                
                if calibri_paths and calibrib_paths:
                    pdfmetrics.registerFont(TTFont('GoodFont', calibri_paths[0]))
                    pdfmetrics.registerFont(TTFont('GoodFontBold', calibrib_paths[0]))
                    good_fonts_available = True
                else:
                    # Fallback to Arial
                    arial_paths = glob.glob("C:/Windows/Fonts/arial*.ttf")
                    arialbd_paths = glob.glob("C:/Windows/Fonts/arialbd*.ttf")
                    
                    if arial_paths and arialbd_paths:
                        pdfmetrics.registerFont(TTFont('GoodFont', arial_paths[0]))
                        pdfmetrics.registerFont(TTFont('GoodFontBold', arialbd_paths[0]))
                        good_fonts_available = True
        except Exception as e:
            print(f"Font registration failed: {e}")
            good_fonts_available = False
        
        # Create the final ID card using your template's natural dimensions
        final_buffer = BytesIO()
        
        # Use the specific template file
        template_path = os.path.join(settings.MEDIA_ROOT, 'id_card_templates', 'yellow Minimalist Id Card.png')
        if os.path.exists(template_path):
            # Get template dimensions from the image
            from PIL import Image as PILImage
            template_img = PILImage.open(template_path)
            page_width, page_height = template_img.size
            pagesize = (page_width, page_height)
            
            # Create canvas with template dimensions
            c = canvas.Canvas(final_buffer, pagesize=pagesize)
            
            # Use the uploaded template as background
            c.drawImage(template_path, 0, 0, width=page_width, height=page_height)
        else:
            # Fallback to A4 size if template not found
            pagesize = A4
            page_width, page_height = pagesize
            c = canvas.Canvas(final_buffer, pagesize=pagesize)
            c.setFillColor(colors.Color(0.98, 0.96, 0.92))  # Light beige
            c.rect(0, 0, page_width, page_height, fill=1)
        
        # ---- MARGINS / SCALE ----
        M = int(min(page_width, page_height) * 0.05)
        left, right = M, page_width - M
        top, bottom = page_height - M, M
        
        # Center the content area - move text towards center
        content_width = int(page_width * 0.8)  # Use 80% of page width for content (increased for long emails)
        content_start = (page_width - content_width) // 2  # Center horizontally

        # Consistent font sizes for clean look
        fs_text   = int(page_height * 0.028)   # same size for both labels and values
        lh_row    = int(fs_text * 1.2)         # consistent line height

        # ---------- PHOTO POSITION (moved down and made bigger) ----------
        y = top - int(page_height * 0.20)  # Start photo lower down (was 0.15, now 0.08)
        photo_size = int(min(page_width, page_height) * 0.44)  # Increased photo size for better visibility
        photo_x = (page_width - photo_size) // 2
        photo_y = y - photo_size

        # no border, just the image
        if getattr(customer, "photo", None):
            try:
                c.drawImage(
                    customer.photo.path,
                    photo_x, photo_y,
                    width=photo_size, height=photo_size,
                    preserveAspectRatio=True, anchor='c'
                )
            except Exception:
                pass

        # gap after photo - adjusted for new photo position
        y = photo_y - int(lh_row * 1.0)

        # ---------- INFO BLOCK (TIGHT LABEL:VALUE) ----------
        # Use English labels with clean formatting
        items = [
            ("Name:",              getattr(customer, "name", "") or ""),
            ("Email:",             getattr(customer, "email", "") or ""),
            ("Mobile:",            getattr(customer, "mobile", "") or ""),
            ("Date of Birth:",     getattr(customer, "dob_str", "") or "05.05.1990"),
            ("Blood Group:",       getattr(customer, "blood_group", "") or "Not Set"),
            ("State:",             getattr(customer, "home_state", "") or "Not Set"),
            ("District:",          getattr(customer, "home_district", "") or "Not Set"),
            ("Registration Date:", customer.created_at.strftime("%d.%m.%Y") if getattr(customer, "created_at", None) else "")
        ]

        # Use good fonts - labels bold, values regular for professional look
        label_font = "GoodFontBold" if good_fonts_available else "Helvetica-Bold"  # Labels bold
        value_font = "GoodFont" if good_fonts_available else "Helvetica"
        
        c.setFont(label_font, fs_text)
        # Calculate exact width for each label and add minimal space
        label_x = content_start  # Start from center position instead of left margin
        gutter = 5  # Fixed small gap in pixels, not percentage

        # Helper function for smart text wrapping
        def _wrap_value_lines(c, text, font, size, max_w, max_lines=2):
            """Greedy wrap for long values (works for emails & words). Returns list of lines."""
            c.setFont(font, size)
            if c.stringWidth(text, font, size) <= max_w:
                return [text]

            # Try splitting by space; if no spaces (e.g., emails), fall back to char-split
            parts = text.split(" ")
            if len(parts) == 1:
                # char-based split
                out, line = [], ""
                for ch in text:
                    nxt = line + ch
                    if c.stringWidth(nxt, font, size) > max_w:
                        out.append(line)
                        line = ch
                        if len(out) >= max_lines - 1:  # last line with ellipsis
                            ell = "…"
                            while c.stringWidth(line + ell, font, size) > max_w and len(line) > 0:
                                line = line[:-1]
                            out.append(line + ell)
                            return out
                    else:
                        line = nxt
                if line:
                    out.append(line)
                return out[:max_lines]

            # word-based split
            out, line = [], ""
            for w in parts:
                trial = (line + " " + w).strip()
                if c.stringWidth(trial, font, size) <= max_w:
                    line = trial
                else:
                    out.append(line)
                    line = w
                    if len(out) >= max_lines - 1:
                        # last line + ellipsis if needed
                        ell = "…"
                        while c.stringWidth(line + ell, font, size) > max_w and len(line) > 0:
                            line = line[:-1]
                        out.append(line + ell)
                        return out
            if line:
                out.append(line)
            return out[:max_lines]

        for lbl, val in items:
            # draw label (bold for professional look)
            c.setFont(label_font, fs_text)
            c.drawString(label_x, y, lbl)

            # calculate value position for this specific label
            c.setFont(label_font, fs_text)
            label_width = c.stringWidth(lbl, label_font, fs_text)
            value_x = label_x + label_width + gutter  # individual calculation for each label
            value_max_w = content_start + content_width - value_x  # Use content area boundary

            # draw value (regular weight, same font size)
            c.setFont(value_font, fs_text)
            
            # Special handling for email - always keep in single line
            if lbl == "Email:":
                # Check if email fits in single line, if not reduce font size
                email_width = c.stringWidth(str(val), value_font, fs_text)
                if email_width > value_max_w:
                    # Calculate reduced font size to fit email in single line
                    reduced_font_size = int(fs_text * (value_max_w / email_width) * 0.95)  # 95% for safety margin
                    c.setFont(value_font, reduced_font_size)
                    lines = [str(val)]  # Single line
                else:
                    lines = [str(val)]  # Single line with normal font
            else:
                lines = _wrap_value_lines(c, str(val), value_font, fs_text, value_max_w, max_lines=1)
            for i, line in enumerate(lines):
                c.drawString(value_x, y - i * lh_row, line)

            # per-field spacing: if we wrapped email in 2 lines, drop more; otherwise normal
            used_lines = len(lines)
            y -= lh_row * used_lines

            # better spacing between different fields for professional look
            y -= int(lh_row * 0.25)

                # ---------- MEMBER ID SECTION ----------
        # Add Member ID after all other fields with some spacing
        y -= int(lh_row * 0.5)  # Small gap after last field
        
        # Member ID section
        c.setFont(label_font, fs_text)
        member_id_label = "Member ID:"
        c.drawString(label_x, y, member_id_label)
        
        # Calculate value position for Member ID
        label_width = c.stringWidth(member_id_label, label_font, fs_text)
        member_value_x = label_x + label_width + gutter
        
        c.setFont(value_font, fs_text)
        c.drawString(member_value_x, y, customer.md_code)
        
        c.save()
        
        # Save to file
        filename = f"ID_Card_{customer.md_code}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, 'id_cards', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Write to file
        with open(filepath, 'wb') as f:
            f.write(final_buffer.getvalue())
        
        final_buffer.close()
        
        return True
        
    except Exception as e:
        print(f"Error generating ID card for {customer.name}: {e}")
        return False

@staff_member_required
def admin_download_id_card(request, customer_id):
    """Admin view to download a user's ID card"""
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        messages.error(request, 'Customer not found.')
        return redirect('admin_id_card_management')
    
    # Generate ID card if it doesn't exist
    filename = f"ID_Card_{customer.md_code}.pdf"
    filepath = os.path.join(settings.MEDIA_ROOT, 'id_cards', filename)
    
    if not os.path.exists(filepath):
        generate_user_id_card(customer)
    
    # Check if file exists after generation
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    else:
        messages.error(request, 'Failed to generate ID card.')
        return redirect('admin_id_card_management')



